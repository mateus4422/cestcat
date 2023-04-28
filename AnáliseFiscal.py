import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from openpyxl import load_workbook

# Título da aplicação
st.title("Análise Fiscal")

# Solicitar o link raw do arquivo do GitHub
url = st.text_input("https://github.com/mateus4422/cestcat/raw/cestcat/Cabecalho_Analise.xlsx"

def analisefiscal():
    def load_xlsx_from_url(url):
        response = requests.get(url)
        content = response.content
        wb = load_workbook(BytesIO(content), read_only=True)
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        data = ws.values
        columns = next(data)[0:]
        df = pd.DataFrame(data, columns=columns)
        return df
    
    # Carregar e exibir a planilha xlsx
    if url:
        try:
            df = load_xlsx_from_url(url)
            st.write(df)
        except Exception as e:
            st.error(f"Erro ao carregar o arquivo xlsx: {e}")
